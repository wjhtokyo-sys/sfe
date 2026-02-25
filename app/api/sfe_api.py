from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect, text
from sqlalchemy.orm import Session

from app.core.auth import get_current_user, require_roles
from app.core.database import get_db
from app.models.entities import Allocation, Bill, BillLine, Customer, CustomerOrder, FifoPendingTask, InventoryLot, Item, PurchaseOrder, PurchaseOrderLine, Supplier, User
from app.schemas.sfe import AllocateIn, BuildBillIn, CustomerIn, ItemIn, LotIn, OrderIn, StateActionIn
from app.services import sfe_service

router = APIRouter(prefix="/api", tags=["SFE"])


@router.post("/customers")
def create_customer(payload: CustomerIn, db: Session = Depends(get_db), _=Depends(require_roles('admin', 'super_admin'))):
    return sfe_service.create_customer(db, payload.name)


@router.get("/customers")
def list_customers(db: Session = Depends(get_db), _=Depends(require_roles('admin', 'super_admin'))):
    return db.query(Customer).order_by(Customer.id.desc()).all()


@router.post('/super/customers')
def super_create_customer_user(payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    username = str(payload.get('username', '')).strip()
    password = str(payload.get('password', '')).strip()
    customer_name = str(payload.get('customer_name', '')).strip()
    if not username or not password or not customer_name:
        raise HTTPException(400, '请填写完整客户账号信息')
    if db.query(User).filter(User.username == username).first():
        raise HTTPException(400, '用户名已存在')
    c = Customer(name=customer_name, is_active=True)
    db.add(c)
    db.flush()
    u = User(username=username, password=password, role='customer', customer_id=c.id, is_active=True)
    db.add(u)
    db.commit()
    return {'ok': True}


@router.get('/super/customers')
def super_customer_users(db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    users = db.query(User).filter(User.role == 'customer').order_by(User.id.desc()).all()
    out = []
    for u in users:
        c = db.get(Customer, u.customer_id) if u.customer_id else None
        out.append({
            'user_id': u.id,
            'username': u.username,
            'is_active': u.is_active,
            'customer_id': u.customer_id,
            'customer_name': c.name if c else None,
            'customer_active': c.is_active if c else None,
        })
    return out


@router.patch('/super/customers/{user_id}')
def super_update_customer_user(user_id: int, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    u = db.get(User, user_id)
    if not u or u.role != 'customer':
        raise HTTPException(404, '客户用户不存在')

    new_username = str(payload.get('username', '')).strip()
    if new_username and new_username != u.username:
        exists = db.query(User).filter(User.username == new_username, User.id != user_id).first()
        if exists:
            raise HTTPException(400, '用户名已存在')
        u.username = new_username

    if 'password' in payload and payload['password']:
        u.password = payload['password']
    if 'is_active' in payload:
        u.is_active = bool(payload['is_active'])

    if u.customer_id:
        c = db.get(Customer, u.customer_id)
        if c and 'is_active' in payload:
            c.is_active = bool(payload['is_active'])
        if c and payload.get('customer_name'):
            c.name = str(payload['customer_name']).strip()

    db.commit()
    return {'ok': True}


@router.delete('/super/customers/{user_id}')
def super_delete_customer_user(user_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    u = db.get(User, user_id)
    if not u or u.role != 'customer':
        raise HTTPException(404, '客户用户不存在')
    c = db.get(Customer, u.customer_id) if u.customer_id else None
    db.delete(u)
    if c:
        db.delete(c)
    db.commit()
    return {'ok': True}


@router.patch('/super/customers/customer/{customer_id}')
def super_update_customer_only(customer_id: int, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    c = db.get(Customer, customer_id)
    if not c:
        raise HTTPException(404, '客户不存在')
    if payload.get('customer_name'):
        c.name = str(payload.get('customer_name')).strip()
    if 'is_active' in payload:
        c.is_active = bool(payload.get('is_active'))
    db.commit()
    return {'ok': True}


@router.delete('/super/customers/customer/{customer_id}')
def super_delete_customer_only(customer_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    c = db.get(Customer, customer_id)
    if not c:
        raise HTTPException(404, '客户不存在')
    has_user = db.query(User).filter(User.customer_id == customer_id).first()
    if has_user:
        raise HTTPException(400, '该客户存在账号，请先删除账号')
    has_order = db.query(CustomerOrder).filter(CustomerOrder.customer_id == customer_id).first()
    has_bill = db.query(Bill).filter(Bill.customer_id == customer_id).first()
    if has_order or has_bill:
        raise HTTPException(400, '该客户已有业务数据，禁止删除')
    db.delete(c)
    db.commit()
    return {'ok': True}


@router.get('/super/admin-users')
def super_admin_users(db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    users = db.query(User).filter(User.role.in_(['admin', 'super_admin'])).order_by(User.id.desc()).all()
    return [{'user_id': u.id, 'username': u.username, 'role': u.role, 'is_active': u.is_active} for u in users]


@router.post('/super/admin-users')
def super_create_admin_user(payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    username = str(payload.get('username', '')).strip()
    password = str(payload.get('password', '')).strip()
    role = str(payload.get('role', 'admin')).strip()
    if role not in {'admin', 'super_admin'}:
        raise HTTPException(400, '角色仅支持admin或super_admin')
    if not username or not password:
        raise HTTPException(400, '请填写用户名和密码')
    if db.query(User).filter(User.username == username).first():
        raise HTTPException(400, '用户名已存在')
    db.add(User(username=username, password=password, role=role, is_active=True))
    db.commit()
    return {'ok': True}


@router.patch('/super/admin-users/{user_id}')
def super_update_admin_user(user_id: int, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    u = db.get(User, user_id)
    if not u or u.role not in {'admin', 'super_admin'}:
        raise HTTPException(404, '管理账号不存在')
    if payload.get('username'):
        new_username = str(payload.get('username')).strip()
        exists = db.query(User).filter(User.username == new_username, User.id != user_id).first()
        if exists:
            raise HTTPException(400, '用户名已存在')
        u.username = new_username
    if payload.get('password'):
        u.password = str(payload.get('password')).strip()
    if payload.get('role') in {'admin', 'super_admin'}:
        u.role = payload.get('role')
    if 'is_active' in payload:
        u.is_active = bool(payload.get('is_active'))
    db.commit()
    return {'ok': True}


@router.delete('/super/admin-users/{user_id}')
def super_delete_admin_user(user_id: int, db: Session = Depends(get_db), me=Depends(get_current_user)):
    if me.role != 'super_admin':
        raise HTTPException(403, '没有权限')
    if me.id == user_id:
        raise HTTPException(400, '不能删除当前登录账号')
    u = db.get(User, user_id)
    if not u or u.role not in {'admin', 'super_admin'}:
        raise HTTPException(404, '管理账号不存在')
    if u.role == 'super_admin':
        cnt = db.query(User).filter(User.role == 'super_admin', User.is_active == True, User.id != user_id).count()
        if cnt < 1:
            raise HTTPException(400, '至少保留一个可用超级管理员账号')
    db.delete(u)
    db.commit()
    return {'ok': True}


@router.get('/suppliers')
def list_suppliers(db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    return db.query(Supplier).order_by(Supplier.id.desc()).all()


@router.post('/suppliers')
def create_supplier(payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    code = str(payload.get('supplier_code', '')).strip()
    name = str(payload.get('name', '')).strip()
    if not code or not name:
        raise HTTPException(400, '请填写供应商编码和供应商名')
    if db.query(Supplier).filter(Supplier.supplier_code == code).first():
        raise HTTPException(400, '供应商编码已存在')
    obj = Supplier(supplier_code=code, name=name, is_active=True)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.patch('/suppliers/{supplier_id}')
def update_supplier(supplier_id: int, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    obj = db.get(Supplier, supplier_id)
    if not obj:
        raise HTTPException(404, '供应商不存在')
    if payload.get('supplier_code'):
        code = str(payload.get('supplier_code')).strip()
        exists = db.query(Supplier).filter(Supplier.supplier_code == code, Supplier.id != supplier_id).first()
        if exists:
            raise HTTPException(400, '供应商编码已存在')
        obj.supplier_code = code
    if payload.get('name'):
        obj.name = str(payload.get('name')).strip()
    if 'is_active' in payload:
        obj.is_active = bool(payload.get('is_active'))
    db.commit()
    return {'ok': True}


@router.delete('/suppliers/{supplier_id}')
def delete_supplier(supplier_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    obj = db.get(Supplier, supplier_id)
    if not obj:
        raise HTTPException(404, '供应商不存在')
    db.delete(obj)
    db.commit()
    return {'ok': True}


@router.get('/purchase-orders/import-template')
def purchase_order_import_template(_=Depends(require_roles('super_admin'))):
    wb = Workbook()
    ws = wb.active
    ws.title = 'purchase_order'
    ws.append(['jan', 'item_name', 'qty', 'unit_cost', 'payment_status', 'purchased_at'])
    ws.append(['JAN00001', '示例货品', 10, 120, 'unpaid', '2026-02-25'])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=purchase_order_template.xlsx'})


@router.post('/purchase-orders/import-excel')
def import_purchase_order_excel(supplier_id: int = Form(...), file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    supplier = db.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(400, '供应商不存在，请先选择供应商')

    try:
        wb = load_workbook(filename=BytesIO(file.file.read()), data_only=True)
        ws = wb.active
        header = [str(c).strip().lower() if c is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        if not rows:
            raise HTTPException(400, '模板没有可导入数据')

        po_no = f"PO{datetime.now().strftime('%Y%m%d%H%M%S')}{supplier.supplier_code}"
        while db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no).first():
            po_no = f"PO{datetime.now().strftime('%Y%m%d%H%M%S%f')}{supplier.supplier_code}"

        # 支持两种模板：
        # A) purchase模板: jan,item_name,qty,unit_cost,payment_status,purchased_at
        # B) item模板兼容: jan,brand,name,spec,msrp_price,in_qty,is_active
        if header[:6] == ['jan', 'item_name', 'qty', 'unit_cost', 'payment_status', 'purchased_at']:
            payment_status = str(rows[0][4] or 'unpaid')
            purchased_at = rows[0][5] if isinstance(rows[0][5], datetime) else datetime.now()
            parsed = [
                {
                    'jan': str(r[0]).strip(),
                    'item_name': str(r[1]).strip(),
                    'qty': int(float(r[2] or 0)),
                    'unit_cost': float(r[3] or 0),
                }
                for r in rows
            ]
        else:
            payment_status = 'unpaid'
            purchased_at = datetime.now()
            parsed = [
                {
                    'jan': str(r[0]).strip(),
                    'item_name': str((r[2] if len(r) > 2 else '') or r[0]).strip(),
                    'qty': int(float((r[5] if len(r) > 5 else 0) or 0)),
                    'unit_cost': float((r[4] if len(r) > 4 else 0) or 0),
                }
                for r in rows
            ]

        po = PurchaseOrder(po_no=po_no, supplier_id=supplier_id, payment_status=payment_status, status='created_unchecked', purchased_at=purchased_at)
        db.add(po)
        db.flush()
        total = 0.0
        for p in parsed:
            line_total = p['qty'] * p['unit_cost']
            total += line_total
            db.add(PurchaseOrderLine(purchase_order_id=po.id, jan=p['jan'], item_name_snapshot=p['item_name'], qty=p['qty'], unit_cost=p['unit_cost'], line_total=line_total))
        po.total_cost = total
        db.commit()
        return {'ok': True, 'po_no': po_no, 'rows': len(parsed)}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(400, f'导入失败：{e}')


@router.get('/purchase-orders')
def list_purchase_orders(db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    return db.query(PurchaseOrder).order_by(PurchaseOrder.purchased_at.desc(), PurchaseOrder.id.desc()).all()


@router.get('/purchase-orders/{po_id}/lines')
def list_purchase_order_lines(po_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    po = db.get(PurchaseOrder, po_id)
    if not po:
        raise HTTPException(404, '进货单不存在')
    return db.query(PurchaseOrderLine).filter(PurchaseOrderLine.purchase_order_id == po_id).order_by(PurchaseOrderLine.id.asc()).all()


@router.delete('/purchase-order-lines/{line_id}')
def delete_purchase_order_line(line_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    line = db.get(PurchaseOrderLine, line_id)
    if not line:
        raise HTTPException(404, '到货明细不存在')
    po = db.get(PurchaseOrder, line.purchase_order_id)
    if po and po.status == 'checked_inbound':
        raise HTTPException(400, '已盘点入库的到货明细不可删除')

    db.query(FifoPendingTask).filter(FifoPendingTask.purchase_order_line_id == line_id).delete()
    db.delete(line)

    if po:
        remain = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.purchase_order_id == po.id).all()
        po.total_cost = float(sum((r.line_total or 0) for r in remain))
        if not remain:
            db.delete(po)

    db.commit()
    return {'ok': True}


@router.get('/arrival-overview')
def arrival_overview(db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    rows = (
        db.query(PurchaseOrderLine, PurchaseOrder)
        .join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderLine.purchase_order_id)
        .order_by(PurchaseOrder.purchased_at.desc(), PurchaseOrderLine.id.asc())
        .all()
    )
    out = []
    for ln, po in rows:
        item = db.query(Item).filter(Item.jan == ln.jan).first()
        customer_name = ''
        alloc_qty = 0
        if item:
            allocs = db.query(Allocation).filter(Allocation.item_id == item.id, Allocation.allocated_by == f'purchase_checkin:{po.po_no}').all()
            if allocs:
                cids = sorted({a.customer_id for a in allocs})
                names = [db.get(Customer, cid).name for cid in cids if db.get(Customer, cid)]
                customer_name = ' / '.join(names)
                alloc_qty = int(sum(a.qty_allocated or 0 for a in allocs))

        out.append({
            'line_id': ln.id,
            'po_no': po.po_no,
            'purchased_at': po.purchased_at,
            'jan': ln.jan,
            'item_name': ln.item_name_snapshot,
            'qty': alloc_qty,
            'unit_cost': ln.unit_cost,
            'customer_name': customer_name,
        })
    return out


@router.get('/arrival-bill-candidates')
def arrival_bill_candidates(customer_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    allocs = (
        db.query(Allocation)
        .filter(Allocation.customer_id == customer_id, Allocation.status == 'active', Allocation.allocated_by.like('purchase_checkin:%'))
        .order_by(Allocation.id.desc())
        .all()
    )
    out = []
    for a in allocs:
        order = db.get(CustomerOrder, a.order_line_id)
        item = db.get(Item, a.item_id)
        po_no = (a.allocated_by or '').split(':', 1)[1] if ':' in (a.allocated_by or '') else ''
        po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no).first() if po_no else None
        pol = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.purchase_order_id == po.id, PurchaseOrderLine.jan == (item.jan if item else '')).first() if po else None
        out.append({
            'allocation_id': a.id,
            'po_no': po_no,
            'purchased_at': po.purchased_at if po else None,
            'jan': item.jan if item else (order.jan_snapshot if order else ''),
            'item_name': item.name if item else (order.item_name_snapshot if order else ''),
            'qty': a.qty_allocated,
            'purchase_unit_price': pol.unit_cost if pol else 0,
            'order_date': order.created_at if order else None,
        })
    return out


@router.post('/bills/from-arrival')
def build_bill_from_arrival(payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    customer_id = int(payload.get('customer_id') or 0)
    lines = payload.get('lines') or []
    if not customer_id or not isinstance(lines, list) or not lines:
        raise HTTPException(400, '参数不完整')

    alloc_ids = [int(x.get('allocation_id')) for x in lines if x.get('allocation_id')]
    allocs = db.query(Allocation).filter(Allocation.id.in_(alloc_ids), Allocation.customer_id == customer_id, Allocation.status == 'active').all()
    if len(allocs) != len(alloc_ids):
        raise HTTPException(400, '存在不可用分配记录')

    bill_no = f"B{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    bill = Bill(customer_id=customer_id, bill_no=bill_no, status='issued')
    db.add(bill)
    db.flush()

    total = 0.0
    price_map = {int(x['allocation_id']): float(x.get('sale_unit_price') or 0) for x in lines if x.get('allocation_id')}
    for a in allocs:
        order = db.get(CustomerOrder, a.order_line_id)
        price = price_map.get(a.id, 0)
        if price <= 0:
            raise HTTPException(400, '销售价格必须大于0')
        line_amount = a.qty_allocated * price
        total += line_amount
        db.add(BillLine(
            bill_id=bill.id,
            allocation_id=a.id,
            item_id=a.item_id,
            jan_snapshot=order.jan_snapshot if order else '',
            item_name_snapshot=order.item_name_snapshot if order else '',
            qty=a.qty_allocated,
            sale_unit_price=price,
            line_amount=line_amount,
        ))
        a.status = 'billed'

    bill.total_amount = total
    db.commit()
    return {'ok': True, 'bill_no': bill_no, 'total_amount': total}


@router.post('/purchase-orders/lines')
def add_purchase_order_line(payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    supplier_id = int(payload.get('supplier_id') or 0)
    supplier = db.get(Supplier, supplier_id) if supplier_id else None
    if not supplier:
        raise HTTPException(400, '供应商名必填')

    rows = payload.get('lines') if isinstance(payload.get('lines'), list) else [payload]
    parsed = []
    for r in rows:
        jan = str(r.get('jan', '')).strip()
        item_name = str(r.get('item_name', '')).strip()
        qty = r.get('qty')
        unit_cost = r.get('unit_cost')
        if not jan or not item_name or qty in (None, '') or unit_cost in (None, ''):
            continue
        parsed.append({
            'jan': jan,
            'item_name': item_name,
            'qty': int(qty),
            'unit_cost': float(unit_cost),
        })

    if not parsed:
        raise HTTPException(400, '至少填写一行完整的到货信息')

    po_no = f"PO{datetime.now().strftime('%Y%m%d%H%M%S%f')}{supplier.supplier_code}"
    while db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no).first():
        po_no = f"PO{datetime.now().strftime('%Y%m%d%H%M%S%f')}{supplier.supplier_code}"

    po = PurchaseOrder(
        po_no=po_no,
        supplier_id=supplier.id,
        payment_status=str(payload.get('payment_status') or 'unpaid'),
        status='created_unchecked',
        purchased_at=datetime.now(),
    )
    db.add(po)
    db.flush()

    total = 0.0
    for p in parsed:
        line_total = p['qty'] * p['unit_cost']
        total += line_total
        db.add(PurchaseOrderLine(
            purchase_order_id=po.id,
            jan=p['jan'],
            item_name_snapshot=p['item_name'],
            qty=p['qty'],
            unit_cost=p['unit_cost'],
            line_total=line_total,
        ))

    po.total_cost = total
    db.commit()
    return {'ok': True, 'po_no': po_no, 'rows': len(parsed)}


@router.delete('/purchase-orders/{po_id}')
def delete_purchase_order(po_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    po = db.get(PurchaseOrder, po_id)
    if not po:
        raise HTTPException(404, '进货单不存在')
    lines = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.purchase_order_id == po.id).all()
    for ln in lines:
        db.delete(ln)
    db.delete(po)
    db.commit()
    return {'ok': True}


@router.post('/purchase-orders/{po_id}/status')
def update_purchase_order_status(po_id: int, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    po = db.get(PurchaseOrder, po_id)
    if not po:
        raise HTTPException(404, '进货单不存在')
    new_status = str(payload.get('status', '')).strip()
    if new_status not in {'created_unchecked', 'checked_inbound'}:
        raise HTTPException(400, '状态不支持')
    if po.status == 'checked_inbound':
        raise HTTPException(400, '进货单已盘点入库，不能回退')

    if new_status == 'checked_inbound':
        lines = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.purchase_order_id == po.id).all()
        for ln in lines:
            item = db.query(Item).filter(Item.jan == ln.jan).first()
            if not item:
                item = Item(jan=ln.jan, brand='-', name=ln.item_name_snapshot or ln.jan, spec=None, msrp_price=0, in_qty=1, is_active=True)
                db.add(item)
                db.flush()

            open_orders = (
                db.query(CustomerOrder)
                .filter(CustomerOrder.item_id == item.id, CustomerOrder.status == 'open')
                .order_by(CustomerOrder.created_at.asc(), CustomerOrder.id.asc())
                .all()
            )
            customer_ids = sorted({o.customer_id for o in open_orders})

            if len(customer_ids) >= 2:
                exists = db.query(FifoPendingTask).filter(
                    FifoPendingTask.purchase_order_line_id == ln.id,
                    FifoPendingTask.reason_code == 'multi_customer_match',
                    FifoPendingTask.status == 'pending'
                ).first()
                if not exists:
                    db.add(FifoPendingTask(
                        purchase_order_line_id=ln.id,
                        source_po_no=po.po_no,
                        jan=ln.jan,
                        item_name=ln.item_name_snapshot,
                        qty=ln.qty,
                        reason_code='multi_customer_match',
                        reason_text='JAN命中多个客户未完成订单，需人工干预',
                        status='pending',
                    ))
                continue

            if len(customer_ids) == 0:
                exists = db.query(FifoPendingTask).filter(
                    FifoPendingTask.purchase_order_line_id == ln.id,
                    FifoPendingTask.reason_code == 'no_order_match',
                    FifoPendingTask.status == 'pending'
                ).first()
                if not exists:
                    db.add(FifoPendingTask(
                        purchase_order_line_id=ln.id,
                        source_po_no=po.po_no,
                        jan=ln.jan,
                        item_name=ln.item_name_snapshot,
                        qty=ln.qty,
                        reason_code='no_order_match',
                        reason_text='未命中客户订单，需人工处理',
                        status='pending',
                    ))
                continue

            total_need = 0
            for od in open_orders:
                need = (od.qty_requested or 0) - (od.qty_allocated or 0)
                if need > 0:
                    total_need += need

            alloc_qty = min(ln.qty, total_need)
            overflow_qty = max(0, ln.qty - alloc_qty)

            if alloc_qty > 0:
                next_rank = db.query(InventoryLot).filter(InventoryLot.item_id == item.id).count() + 1
                lot = InventoryLot(item_id=item.id, qty_received=alloc_qty, qty_remaining=alloc_qty, fifo_rank=next_rank, location='PO_CHECKED')
                db.add(lot)
                db.flush()

                remaining = lot.qty_remaining
                for od in open_orders:
                    if remaining <= 0:
                        break
                    need = (od.qty_requested or 0) - (od.qty_allocated or 0)
                    if need <= 0:
                        continue
                    take = min(remaining, need)
                    alloc = Allocation(
                        customer_id=od.customer_id,
                        order_line_id=od.id,
                        lot_id=lot.id,
                        item_id=item.id,
                        qty_allocated=take,
                        fifo_rank_snapshot=lot.fifo_rank,
                        status='active',
                        allocated_by=f'purchase_checkin:{po.po_no}',
                    )
                    db.add(alloc)
                    od.qty_allocated = (od.qty_allocated or 0) + take
                    od.status = 'closed' if od.qty_allocated >= od.qty_requested else 'open'
                    remaining -= take

                lot.qty_remaining = remaining

            if overflow_qty > 0:
                exists = db.query(FifoPendingTask).filter(
                    FifoPendingTask.purchase_order_line_id == ln.id,
                    FifoPendingTask.reason_code == 'no_order_match',
                    FifoPendingTask.status == 'pending'
                ).first()
                if not exists:
                    db.add(FifoPendingTask(
                        purchase_order_line_id=ln.id,
                        source_po_no=po.po_no,
                        jan=ln.jan,
                        item_name=ln.item_name_snapshot,
                        qty=overflow_qty,
                        reason_code='no_order_match',
                        reason_text='单客户命中后超出订单剩余数量，超出部分需人工处理',
                        status='pending',
                    ))

        po.status = 'checked_inbound'
    else:
        po.status = new_status

    db.commit()
    return {'ok': True}


@router.post('/purchase-orders/{po_id}/pay')
def mark_purchase_order_paid(po_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    po = db.get(PurchaseOrder, po_id)
    if not po:
        raise HTTPException(404, '进货单不存在')
    po.payment_status = 'paid'
    db.commit()
    return {'ok': True}


@router.post('/purchase-orders/reset-status')
def reset_purchase_orders_status(payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请先选择进货单')
    rows = db.query(PurchaseOrder).filter(PurchaseOrder.id.in_(ids)).all()
    for po in rows:
        po.payment_status = 'unpaid'
        po.status = 'created_unchecked'
    db.commit()
    return {'ok': True, 'count': len(rows)}


@router.get('/fifo/pending')
def list_fifo_pending(db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    tasks = db.query(FifoPendingTask).order_by(FifoPendingTask.id.desc()).all()
    out = []
    for t in tasks:
        ln = db.get(PurchaseOrderLine, t.purchase_order_line_id) if t.purchase_order_line_id else None
        po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == t.source_po_no).first()
        out.append({
            'id': t.id,
            'source_po_no': t.source_po_no,
            'purchased_at': po.purchased_at if po else None,
            'jan': t.jan,
            'item_name': t.item_name,
            'qty': t.qty,
            'unit_cost': ln.unit_cost if ln else None,
            'reason_code': t.reason_code,
            'reason_text': t.reason_text,
            'status': t.status,
        })
    return out


@router.delete('/fifo/pending/{task_id}')
def delete_fifo_pending_task(task_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    task = db.get(FifoPendingTask, task_id)
    if not task:
        raise HTTPException(404, '挂起任务不存在')
    db.delete(task)
    db.commit()
    return {'ok': True}


@router.post('/fifo/pending/{task_id}/resolve')
def resolve_fifo_pending_task(task_id: int, payload: dict, db: Session = Depends(get_db), user=Depends(require_roles('super_admin'))):
    task = db.get(FifoPendingTask, task_id)
    if not task:
        raise HTTPException(404, '挂起任务不存在')
    if task.status == 'resolved':
        raise HTTPException(400, '该任务已处理')
    action = str(payload.get('action', '')).strip()
    if action not in {'inbound_to_order', 'inbound_stock', 'close_only'}:
        raise HTTPException(400, '不支持的处理动作')

    if action in {'inbound_to_order', 'inbound_stock'}:
        line = db.get(PurchaseOrderLine, task.purchase_order_line_id) if task.purchase_order_line_id else None
        if not line:
            raise HTTPException(400, '来源进货明细不存在')
        item = db.query(Item).filter(Item.jan == line.jan).first()
        if not item:
            item = Item(jan=line.jan, brand='-', name=line.item_name_snapshot or line.jan, spec=None, msrp_price=0, in_qty=1, is_active=True)
            db.add(item)
            db.flush()
        lot = InventoryLot(item_id=item.id, qty_received=line.qty, qty_remaining=line.qty, fifo_rank=db.query(InventoryLot).filter(InventoryLot.item_id == item.id).count() + 1, location='FIFO_PENDING')
        db.add(lot)

    task.status = 'resolved'
    task.resolution_note = action
    task.resolved_by = 'super_admin'
    task.resolved_at = __import__('datetime').datetime.utcnow()
    db.commit()
    return {'ok': True}


@router.post("/items")
def create_item(payload: ItemIn, db: Session = Depends(get_db), _=Depends(require_roles('admin', 'super_admin'))):
    return sfe_service.create_item(db, payload.model_dump())


@router.get("/items")
def list_items(keyword: str | None = None, db: Session = Depends(get_db), _=Depends(get_current_user)):
    return sfe_service.search_items(db, keyword)


@router.patch('/items/{item_id}')
def update_item(item_id: int, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    item = db.get(Item, item_id)
    if not item:
        raise HTTPException(404, '商品不存在')
    for k in ['jan', 'brand', 'name', 'spec', 'msrp_price', 'in_qty', 'is_active']:
        if k in payload:
            setattr(item, k, payload[k])
    db.commit()
    return {'ok': True}


@router.delete('/items/{item_id}')
def delete_item(item_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    item = db.get(Item, item_id)
    if not item:
        raise HTTPException(404, '商品不存在')
    db.delete(item)
    db.commit()
    return {'ok': True}


@router.get('/items/import-template')
def item_import_template(_=Depends(require_roles('super_admin'))):
    wb = Workbook()
    ws = wb.active
    ws.title = 'items'
    ws.append(['jan', 'brand', 'name', 'spec', 'msrp_price', 'in_qty', 'is_active'])
    ws.append(['JAN00001', 'Shimano', '示例商品', '规格', 199, 1, 1])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=item_template.xlsx'})


@router.post('/items/import-excel')
def import_items_excel(file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    try:
        raw = file.file.read()
        wb = load_workbook(filename=BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, 'Excel文件解析失败，请使用模板并保存为xlsx后重试')

    def to_float(v, default=0.0):
        if v is None or str(v).strip() == '':
            return default
        return float(v)

    def to_int(v, default=1):
        if v is None or str(v).strip() == '':
            return default
        return int(float(v))

    def to_bool(v, default=True):
        if v is None or str(v).strip() == '':
            return default
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if s in {'1', 'true', 'yes', 'y', '是', '启用'}:
            return True
        if s in {'0', 'false', 'no', 'n', '否', '停用'}:
            return False
        return default

    def normalize_jan(v):
        if v is None:
            return ''
        if isinstance(v, (int, float)):
            return str(int(v))
        s = str(v).strip()
        if s.endswith('.0') and s.replace('.', '', 1).isdigit():
            return s[:-2]
        return s

    created = 0
    updated = 0
    try:
        existing = {normalize_jan(i.jan): i for i in db.query(Item).all()}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            jan, brand, name, spec, msrp_price, in_qty, is_active = (list(row) + [None] * 7)[:7]
            jan = normalize_jan(jan)
            if not jan:
                continue

            obj = existing.get(jan)
            if obj:
                obj.brand = str(brand or obj.brand)
                obj.name = str(name or obj.name)
                obj.spec = str(spec) if spec else None
                obj.msrp_price = to_float(msrp_price, obj.msrp_price or 0)
                obj.in_qty = to_int(in_qty, obj.in_qty or 1)
                obj.is_active = to_bool(is_active, obj.is_active)
                updated += 1
            else:
                obj = Item(
                    jan=jan,
                    brand=str(brand or ''),
                    name=str(name or ''),
                    spec=str(spec) if spec else None,
                    msrp_price=to_float(msrp_price, 0),
                    in_qty=to_int(in_qty, 1),
                    is_active=to_bool(is_active, True),
                )
                db.add(obj)
                existing[jan] = obj
                created += 1
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(400, f'上传失败：{e}')

    return {'created': created, 'updated': updated}


@router.post("/orders")
def create_order(payload: OrderIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if user.role == 'customer' and user.customer_id != payload.customer_id:
        raise HTTPException(403, 'customer can only create own orders')
    return sfe_service.create_order(db, payload.customer_id, payload.item_id, payload.qty_requested)


@router.get("/orders")
def list_orders(db: Session = Depends(get_db), user=Depends(get_current_user)):
    q = db.query(CustomerOrder)
    if user.role == 'customer':
        q = q.filter(CustomerOrder.customer_id == user.customer_id)
    return q.order_by(CustomerOrder.id.desc()).all()


@router.delete('/orders/{order_id}')
def delete_order(order_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    order = db.get(CustomerOrder, order_id)
    if not order:
        raise HTTPException(404, '订单不存在')

    if user.role == 'customer':
        if user.customer_id != order.customer_id:
            raise HTTPException(403, '无权删除该订单')
        if (order.qty_allocated or 0) > 0:
            raise HTTPException(400, '该订单已有分配记录，无法删除')
    elif user.role != 'super_admin':
        raise HTTPException(403, '无权删除订单')

    db.delete(order)
    db.commit()
    return {'ok': True}


@router.get('/orders/{order_id}/arrivals')
def order_arrivals(order_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    order = db.get(CustomerOrder, order_id)
    if not order:
        raise HTTPException(404, '订单不存在')
    if user.role == 'customer' and user.customer_id != order.customer_id:
        raise HTTPException(403, '没有权限')

    allocs = db.query(Allocation).filter(Allocation.order_line_id == order_id).order_by(Allocation.id.asc()).all()
    out = []
    for a in allocs:
        item = db.get(Item, a.item_id)
        arrival_time = None
        if a.allocated_by and ':' in a.allocated_by:
            po_no = a.allocated_by.split(':', 1)[1]
            po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no).first()
            arrival_time = po.purchased_at if po else None
        if arrival_time is None:
            lot = db.get(InventoryLot, a.lot_id)
            arrival_time = lot.created_at if lot else None
        out.append({
            'arrival_time': arrival_time,
            'jan': item.jan if item else order.jan_snapshot,
            'item_name': item.name if item else order.item_name_snapshot,
            'qty': a.qty_allocated,
        })
    return out


@router.post('/bills/from-orders')
def build_bill_from_orders(payload: dict, db: Session = Depends(get_db), user=Depends(require_roles('super_admin'))):
    return sfe_service.build_bill_from_orders(db, payload.get('order_ids', []), user.role, float(payload.get('sale_unit_price', 1)))


@router.post("/lots")
def inbound_lot(payload: LotIn, db: Session = Depends(get_db), _=Depends(require_roles('admin', 'super_admin'))):
    return sfe_service.inbound_lot(db, payload.item_id, payload.qty_received, payload.location)


@router.get("/lots")
def list_lots(db: Session = Depends(get_db), _=Depends(require_roles('admin', 'super_admin'))):
    return db.query(InventoryLot).order_by(InventoryLot.id.desc()).all()


@router.post("/allocations/fifo")
def allocate_fifo(payload: AllocateIn, db: Session = Depends(get_db), _=Depends(require_roles('admin', 'super_admin'))):
    return sfe_service.allocate_fifo(db, payload.order_line_id, payload.allocated_by)


@router.get("/allocations")
def list_allocations(db: Session = Depends(get_db), _=Depends(require_roles('admin', 'super_admin'))):
    return db.query(Allocation).order_by(Allocation.id.desc()).all()


@router.post("/bills")
def build_bill(payload: BuildBillIn, db: Session = Depends(get_db), _=Depends(require_roles('admin', 'super_admin'))):
    return sfe_service.build_bill(db, payload.customer_id, payload.allocation_ids, payload.sale_unit_price)


@router.get("/bills")
def list_bills(db: Session = Depends(get_db), user=Depends(get_current_user)):
    q = db.query(Bill)
    if user.role == 'customer':
        q = q.filter(Bill.customer_id == user.customer_id)
    rows = q.order_by(Bill.id.desc()).all()
    out = []
    for b in rows:
        points = db.query(BillLine).filter(BillLine.bill_id == b.id).with_entities(text('coalesce(sum(qty),0)')).scalar() or 0
        out.append({
            'id': b.id,
            'customer_id': b.customer_id,
            'bill_no': b.bill_no,
            'status': b.status,
            'payment_status': b.payment_status,
            'shipping_status': b.shipping_status,
            'total_amount': b.total_amount,
            'currency': b.currency,
            'created_at': b.created_at,
            'archived_at': b.archived_at,
            'goods_points': int(points),
        })
    return out


@router.get('/bills/{bill_id}/lines')
def bill_lines(bill_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    bill = db.get(Bill, bill_id)
    if not bill:
        raise HTTPException(404, '账单不存在')
    if user.role == 'customer' and user.customer_id != bill.customer_id:
        raise HTTPException(403, '没有权限')

    lines = db.query(BillLine).filter(BillLine.bill_id == bill_id).order_by(BillLine.id.asc()).all()
    out = []
    for bl in lines:
      alloc = db.get(Allocation, bl.allocation_id)
      order = db.get(CustomerOrder, alloc.order_line_id) if alloc else None
      arrival_date = None
      if alloc and alloc.allocated_by and ':' in alloc.allocated_by:
          po_no = alloc.allocated_by.split(':', 1)[1]
          po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no).first()
          arrival_date = po.purchased_at if po else None
      out.append({
          'id': bl.id,
          'jan_snapshot': bl.jan_snapshot,
          'item_name_snapshot': bl.item_name_snapshot,
          'qty': bl.qty,
          'sale_unit_price': bl.sale_unit_price,
          'line_amount': bl.line_amount,
          'order_date': order.created_at if order else None,
          'arrival_date': arrival_date,
      })
    return out


@router.post("/bills/{bill_id}/state")
def bill_state_action(bill_id: int, payload: StateActionIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if user.role not in {'customer', 'super_admin', 'admin'}:
        raise HTTPException(403, '没有权限')
    bill = db.get(Bill, bill_id)
    if not bill:
        raise HTTPException(404, '账单不存在')
    if user.role == 'customer' and user.customer_id != bill.customer_id:
        raise HTTPException(403, '没有权限')
    return sfe_service.update_bill_state(db, bill_id, payload.action, user.role)


@router.patch('/bills/{bill_id}')
def update_bill(bill_id: int, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    bill = db.get(Bill, bill_id)
    if not bill:
        raise HTTPException(404, '账单不存在')

    if 'bill_no' in payload and payload['bill_no']:
        new_no = str(payload['bill_no']).strip()
        exists = db.query(Bill).filter(Bill.bill_no == new_no, Bill.id != bill_id).first()
        if exists:
            raise HTTPException(400, '账单号已存在')
        bill.bill_no = new_no

    for k in ['status', 'payment_status', 'shipping_status', 'currency']:
        if k in payload and payload[k]:
            setattr(bill, k, payload[k])

    if 'total_amount' in payload and payload['total_amount'] is not None:
        bill.total_amount = float(payload['total_amount'])

    db.commit()
    return {'ok': True}


@router.post('/fifo/pending-legacy/{allocation_id}/resolve')
def resolve_fifo_pending(allocation_id: int, decision: StateActionIn, _=Depends(require_roles('super_admin'))):
    return {'allocation_id': allocation_id, 'decision': decision.action, 'status': 'legacy_resolved'}


@router.get('/db/tables')
def db_tables(db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    names = sorted([n for n in inspect(db.bind).get_table_names() if not n.startswith('sqlite_')])
    return {'tables': names}


@router.get('/db/table/{table_name}')
def db_table_rows(table_name: str, limit: int = 100, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    inspector = inspect(db.bind)
    tables = inspector.get_table_names()
    if table_name not in tables:
        raise HTTPException(404, '表不存在')
    safe_limit = max(1, min(limit, 500))
    rows = db.execute(text(f'SELECT * FROM "{table_name}" ORDER BY rowid DESC LIMIT :l'), {'l': safe_limit}).mappings().all()
    cols = [c['name'] for c in inspector.get_columns(table_name)]
    return {'columns': cols, 'rows': [dict(r) for r in rows]}


@router.post('/db/table/{table_name}')
def db_table_insert(table_name: str, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    inspector = inspect(db.bind)
    tables = inspector.get_table_names()
    if table_name not in tables:
        raise HTTPException(404, '表不存在')
    data = payload.get('data') or {}
    if not isinstance(data, dict) or not data:
        raise HTTPException(400, 'data不能为空对象')

    cols = [c['name'] for c in inspector.get_columns(table_name)]
    pks = set(inspector.get_pk_constraint(table_name).get('constrained_columns') or [])
    allowed = [k for k in data.keys() if k in cols and k not in pks]
    if not allowed:
        raise HTTPException(400, '没有可写入字段')

    col_sql = ','.join([f'"{c}"' for c in allowed])
    val_sql = ','.join([f':{c}' for c in allowed])
    db.execute(text(f'INSERT INTO "{table_name}" ({col_sql}) VALUES ({val_sql})'), {c: data[c] for c in allowed})
    db.commit()
    return {'ok': True}


@router.patch('/db/table/{table_name}/{row_id}')
def db_table_update(table_name: str, row_id: int, payload: dict, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    inspector = inspect(db.bind)
    tables = inspector.get_table_names()
    if table_name not in tables:
        raise HTTPException(404, '表不存在')
    data = payload.get('data') or {}
    if not isinstance(data, dict) or not data:
        raise HTTPException(400, 'data不能为空对象')

    pk_cols = inspector.get_pk_constraint(table_name).get('constrained_columns') or []
    if len(pk_cols) != 1:
        raise HTTPException(400, '仅支持单主键表修改')
    pk = pk_cols[0]
    cols = [c['name'] for c in inspector.get_columns(table_name)]
    allowed = [k for k in data.keys() if k in cols and k != pk]
    if not allowed:
        raise HTTPException(400, '没有可修改字段')

    set_sql = ','.join([f'"{c}"=:{c}' for c in allowed])
    params = {c: data[c] for c in allowed}
    params['row_id'] = row_id
    db.execute(text(f'UPDATE "{table_name}" SET {set_sql} WHERE "{pk}"=:row_id'), params)
    db.commit()
    return {'ok': True}


@router.delete('/db/table/{table_name}/{row_id}')
def db_table_delete(table_name: str, row_id: int, db: Session = Depends(get_db), _=Depends(require_roles('super_admin'))):
    inspector = inspect(db.bind)
    tables = inspector.get_table_names()
    if table_name not in tables:
        raise HTTPException(404, '表不存在')

    pk_cols = inspector.get_pk_constraint(table_name).get('constrained_columns') or []
    if len(pk_cols) != 1:
        raise HTTPException(400, '仅支持单主键表删除')
    pk = pk_cols[0]

    for child in tables:
        fks = inspector.get_foreign_keys(child)
        for fk in fks:
            if fk.get('referred_table') == table_name:
                child_col = (fk.get('constrained_columns') or [None])[0]
                parent_col = (fk.get('referred_columns') or [None])[0]
                if child_col and parent_col == pk:
                    cnt = db.execute(text(f'SELECT COUNT(1) c FROM "{child}" WHERE "{child_col}"=:v'), {'v': row_id}).scalar() or 0
                    if cnt > 0:
                        raise HTTPException(400, f'拒绝删除：存在关联数据 {child}({cnt})')

    db.execute(text(f'DELETE FROM "{table_name}" WHERE "{pk}"=:row_id'), {'row_id': row_id})
    db.commit()
    return {'ok': True}
